import numpy as np
import openpyxl
import xlrd
import os.path
import re
import string
try:
    from variable import variable
except ImportError:
    from pyees.variable import variable

def fileFromSheets(Sheets):
    ## TODO save Sheet as xlFile
    raise NotImplementedError()

def sheetsFromFile(xlFile, dataRange: str | list[str], uncertRange: str | list[str] = None, sheets: int | list[int] = None):
    dat = _SheetsFromFile(xlFile, dataRange, uncertRange, sheets)
    if len(dat.dat) == 1:
        return dat.dat[0]
    return dat.dat


class _SheetsFromFile():

    def __init__(self, xlFile, dataRange, uncertRange=None, sheets = None) -> None:


        # check the extension
        extension = os.path.splitext(xlFile)[1]
        supportedExtensions = ['.xls', '.xlsx']
        if extension not in supportedExtensions:
            raise ValueError(f'The file extension is not supported. The supported extension are {supportedExtensions}')

        # parse functions for the specific extension and get all sheets
        if extension == '.xls':
            self.wb = xlrd.open_workbook(xlFile)
            
            self.sheets = self.wb.sheets()
            if not sheets is None:
                if not isinstance(sheets, list):
                    sheets = [sheets]
                self.sheets = [elem for i,elem in enumerate(self.sheets) if i in sheets]

            def readCell(sheet, row, col):
                return sheet.cell(row, col).value

            def readRow(sheet, row):
                return [elem.value for elem in sheet.row(row)]

            def readCol(sheet, col):
                return [elem.value for elem in sheet.col(col)]

        elif extension == '.xlsx':
            self.wb = openpyxl.load_workbook(xlFile, data_only=True)
            self.sheets = [self.wb[elem] for elem in self.wb.sheetnames]
            if not sheets is None:
                if not isinstance(sheets, list):
                    sheets = [sheets]
                self.sheets = [elem for i,elem in enumerate(self.sheets) if i in sheets]
                
            def readCell(sheet, row, col):
                return sheet.cell(row + 1, col + 1).value

            def readRow(sheet, row):
                return [elem.value for elem in list(sheet.iter_rows())[row]]

            def readCol(sheet, col):
                return [elem.value for elem in list(sheet.iter_cols())[col]]

        self.readCell = readCell
        self.readRow = readRow
        self.readCol = readCol

        if not isinstance(dataRange, list):
            dataRange = [dataRange] * len(self.sheets)
        if len(dataRange) != len(self.sheets):
            raise ValueError('The length of the input "dataRange" has to be equal to the length of the intput "sheets"')
        
        if not uncertRange is None:
            if not isinstance(uncertRange, list):
                uncertRange = [uncertRange] * len(self.sheets)
            if len(uncertRange) != len(self.sheets):
                raise ValueError('The length of the input "uncertRange" has to be equal to the length of the intput "sheets"')
        else:
            uncertRange = [None] * len(self.sheets)
            
        
        self.dataStartCol = []
        self.dataEndCol = []
        self.uncertStartCol = []
        self.uncertEndCol = []    
        for i in range(len(self.sheets)):
            if '-' in dataRange[i]:
                index = dataRange[i].find('-')
                dataStartCol = dataRange[i][0:index]
                dataEndCol = dataRange[i][index + 1:]
            else:
                dataStartCol = dataRange[i]
                dataEndCol = dataStartCol

            if '-' in dataStartCol or '-' in dataEndCol:
                raise ValueError('The data range can only include a singly hyphen (-)')

            if not uncertRange[i] is None:
                if '-' in uncertRange[i]:
                    index = uncertRange[i].find('-')
                    uncertStartCol = uncertRange[i][0:index]
                    uncertEndCol = uncertRange[i][index + 1:]
                else:
                    uncertStartCol = uncertRange[i]
                    uncertEndCol = uncertStartCol

                if '-' in uncertStartCol or '-' in uncertEndCol:
                    raise ValueError('The data range can only include a singly hyphen (-)')
            else:
                uncertStartCol = None
                uncertEndCol = None

            # convert the coloumns
            self.dataStartCol.append(self.colToIndex(dataStartCol))
            self.dataEndCol.append(self.colToIndex(dataEndCol))
            self.uncertStartCol.append(self.colToIndex(uncertStartCol))
            self.uncertEndCol.append(self.colToIndex(uncertEndCol))

            # check the number of coloumns
            nColsData = self.dataEndCol[i] - self.dataStartCol[i] + 1
            if not self.uncertStartCol[i] is None:
                nColsUncert = self.uncertEndCol[i] - self.uncertStartCol[i] + 1
                if nColsData != nColsUncert:
                    raise ValueError('The number of coloumns of the data is not equal to the number of coloumns for the uncertanty')
        
        # read the data
        self.readData()

    def colToIndex(self, col):
        if col is None:
            return None
        if not isinstance(col, str):
            raise ValueError('The coloumn has to be a string')
        num = 0
        for c in col:
            if c in string.ascii_letters:
                num = num * 26 + (ord(c.upper()) - ord('A')) + 1
        return num

    def formatHeaders(self, header):
        out = []
        for head in header:
            if head is None:
                out.append('')
                continue
            # remove symbols and replace with _
            head = re.sub(r'[^\w]', '_', head.lower())

            # remove multiple consequtive _
            done = False
            a = head
            while not done:
                b = a.replace('__', '_')
                if a == b:
                    done = True

            # add "_" to the begining of the name if the first letter is a digit
            if head[0].isnumeric():
                head = '_' + head

            # remove "_" if the last letter is "_"
            if head[-1] == "_" and len(head) != 1:
                head = head[0:-1]

            # add a number to the end of the unit if the unit exists
            if head not in out:
                out.append(head)
            else:
                i, imax, done = 1, 100, False
                while not done and i <= imax:
                    h = head + f'_{i+1}'
                    if h not in out:
                        out.append(h)
                        done = True
                    i += 1
                if not done:
                    raise ValueError(f'The header {head} could not be added to the sheet')
        return out

    def readData(self):
        self.dat = []

        # Looping over the sheets in the data file
        for ii, sheet in enumerate(self.sheets):
            
            sheetData = Sheet()
            
            self.nCols = self.dataEndCol[ii] - self.dataStartCol[ii] + 1
            
            # determine the number of variables
            headers = self.readRow(sheet, 0)[self.dataStartCol[ii]-1:self.dataStartCol[ii] - 1 + self.nCols]

            headers = self.formatHeaders(headers)
            units = self.readRow(sheet, 1)[self.dataStartCol[ii]-1:self.dataStartCol[ii] - 1 + self.nCols]

            # determine the number of datapoints
            nDataPoints = []
            for i in range(self.nCols):
                nDataPoint = self.readCol(sheet, self.dataStartCol[ii] + i - 1)[2:]
                nDataPoint = sum([1 if elem not in ['', None] else 0 for elem in nDataPoint])
                nDataPoints.append(nDataPoint)
            if not all(elem == nDataPoints[0] for elem in nDataPoints):
                raise ValueError('There are not an equal amount of rows in the data')
            nDataPoint = nDataPoints[0]

            # read the data
            data = np.zeros([nDataPoint, self.nCols])
            for i in range(nDataPoint):
                for j in range(self.nCols):
                    data[i, j] = float(self.readCell(sheet, 2 + i, j+ self.dataStartCol[ii] - 1))

            if not self.uncertStartCol[ii] is None:
                # determine the number of rows in the uncertanty
                nUncertanties = []
                for i in range(self.nCols):
                    nUncertanty = self.readCol(sheet, self.uncertStartCol[ii] - 1 + i)[2:]
                    nUncertanty = sum([1 if elem not in ['', None] else 0 for elem in nUncertanty])
                    nUncertanties.append(nUncertanty)
                if not all(elem == nUncertanties[0] for elem in nUncertanties):
                    raise ValueError('There are not an equal amount of rows in the uncertanty')
                nUncertanty = nUncertanties[0]

                # evaluate the number of rows of the uncertanty
                if nUncertanty not in [nDataPoint, nDataPoint * self.nCols]:
                    raise ValueError('The number of rows in the uncertanty has to be equal to the number of rows of data or equal to the number of rows of data multiplied with the number of coloumns in the data')

                if nUncertanty == nDataPoint:
                    # There is one row of uncertanty for each row of data. Therefore there are no covariance data in the sheet

                    # read the uncertanty
                    uncert = np.zeros([nDataPoint, self.nCols])
                    for i in range(nDataPoint):
                        for j in range(self.nCols):
                            uncert[i, j] = float(self.readCell(sheet, 2 + i, self.uncertStartCol[ii] - 1 + j))

                    # create the measurements uncertanties
                    for i in range(self.nCols):
                        name = headers[i]
                        unit = units[i]
                        val = np.array(data[:, i])
                        u = np.array(uncert[:, i])
                        var = variable(val, unit, uncert=u)

                        sheetData._addMeasurement(name, var)
                else:
                    # There are covariance data in the sheet

                    # read the uncertanty
                    uncert = []
                    for i in range(nDataPoint):
                        u = np.zeros([self.nCols, self.nCols])
                        for j in range(self.nCols):
                            for k in range(self.nCols):
                                u[j, k] = float(self.readCell(sheet, 2 + i * self.nCols + j, self.uncertStartCol[ii] - 1 + k))
                        uncert.append(u)

                    # check if each element in the uncertanty is symmetric
                    for elem in uncert:
                        if (elem.shape == elem.transpose().shape) and (elem == elem.transpose()).all():
                            pass
                        else:
                            raise ValueError('The covariances has to be symmetric')

                    # create the measurements with covariance uncertanties
                    vars = []
                    for i in range(self.nCols):
                        name = headers[i]
                        unit = units[i]
                        val = np.array(data[:, i])
                        u = np.array([np.sqrt(elem[i, i]) for elem in uncert])
                        var = variable(val, unit, uncert=u)
                        vars.append(var)

                    for i in range(self.nCols):
                        covariance = [elem[:, i] for elem in uncert]
                        for j in range(self.nCols):
                            if i != j:
                                cov = [elem[j] for elem in covariance]
                                vars[i].addCovariance(vars[j], cov, str(vars[i]._unitObject * vars[j]._unitObject))

                    for head, var in zip(headers, vars):
                        sheetData._addMeasurement(head, var)
            else:
                # There are no uncertaty data in the sheet

                # create the measurements without uncertanties
                for i in range(self.nCols):
                    name = headers[i]
                    unit = units[i]
                    val = np.array(data[:, i])
                    var = variable(val, unit)
                    sheetData._addMeasurement(name, var)

            self.dat.append(sheetData)



class Sheet():
    def __init__(self):
        self.measurements = []
        self.measurementNames = []

    def _addMeasurement(self, name, var):
        self.measurements.append(var)
        self.measurementNames.append(name)
        setattr(self, name, var)

    def printContents(self):
        for name in self.measurementNames:
            print(name)

    def __getitem__(self, index):
        measurements = []
        for meas in self.measurements:
            val = meas.value[index]
            unit = meas.unit
            uncert = meas.uncert[index]
            measurements.append(variable(val, unit, uncert))

        sheet = Sheet()

        for measurement, measurementName in zip(measurements, self.measurementNames):
            sheet._addMeasurement(measurementName, measurement)

        return sheet

    def append(self, other):
        if not isinstance(other, Sheet):
            raise ValueError('You can only append two sheets together')

        # Test if all names are the same
        for elem in self.measurementNames:
            if elem not in other.measurementNames:
                raise ValueError('You can only append sheets with the excact same measurements. The names did not match')

        for elem in other.measurementNames:
            if elem not in self.measurementNames:
                raise ValueError('You can only append sheets with the excact same measurements. The names did not match')

        # append the measurements from other to self
        for elem, elemNames in zip(self.measurements, self.measurementNames):
            index = other.measurementNames.index(elemNames)
            elem.append(other.measurements[index])    
        
    def __iter__(self):
        return iter(self.measurements)


    
    


