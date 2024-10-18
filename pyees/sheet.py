from typing import Any, List
import numpy as np
from python_calamine import CalamineWorkbook
import xlwt
import openpyxl
import os.path
import re
import string
try:
    from variable import variable, scalarVariable, arrayVariable, unit
except ImportError:
    from pyees.variable import variable, scalarVariable, arrayVariable, unit



def fileFromSheets(sheets, fileName):
    if not isinstance(sheets, list):
        sheets = [sheets]
    _fileFromSheets(sheets, fileName)

class _fileFromSheets():
    def __init__(self, sheets, fileName) -> None:
        
        self.fileName = fileName
        self.sheets = sheets
        
        extension = os.path.splitext(fileName)[1]
        supportedExtensions = ['.xls', '.xlsx']
        if extension not in supportedExtensions:
            raise ValueError(f'The file extension is not supported. The supported extension are {supportedExtensions}')
            
        if extension == '.xls':
            self.wb = xlwt.Workbook()                            
            
            def createSheet():
                return self.wb.add_sheet()
        
            def write(sheet, row, col, value):
                sheet.write(row, col, value)
            
            def getSheet(index):
                return self.wb.add_sheet(f'Sheet {index + 1}')
                       
            
        elif extension == '.xlsx':
            self.wb = openpyxl.Workbook()
            
            def createSheet():
                return self.wb.create_sheet()
        
            def write(sheet, row, col, value):
                sheet.cell(row+1, col+1, value)
            
            def getSheet(index):
                if index > 0:
                    worksheet = self.createSheet()
                else:
                    worksheet = self.wb.active
                return worksheet
            
        self.createSheet = createSheet
        self.write = write
        self.getSheet = getSheet
        
        self.saveSheets()
        
    def saveSheets(self):
        
        for ii, sheet in enumerate(self.sheets):
            
            worksheet = self.getSheet(ii)
            
            col = 0
            for objectName in dir(sheet):
                object = getattr(sheet, objectName)
                if isinstance(object, scalarVariable):
                    meas = object
 
                    self.write(worksheet, 0, col, objectName)
                    u = '-' if meas.unit == '1' else meas.unit
                    self.write(worksheet, 1, col, u)
                    
                    u = meas._unitObject
                    meas._unitObject = unit('')
                    for elem in meas:
                        elem._unitObject = unit('')
                    
                    for row, val in enumerate(meas):
                        string = str(val)
                        string = string.replace(' +/- ', '±\n')
                        self.write(worksheet, row + 2, col, string)
                    
                    meas._unitObject = u
                    for elem in meas:
                        elem._unitObject = u
                           
                    col += 1          
                         
        self.wb.save(self.fileName)
        
            
            
        
        
        

def sheetsFromFile(xlFile, dataRange: str | List[str], uncertRange: str | List[str] = None, sheets: int | List[int] = None):
    dat = _sheetsFromFile(xlFile, dataRange, uncertRange, sheets)
    if len(dat.dat) == 1:
        return dat.dat[0]
    return dat.dat


class _sheetsFromFile():

    def __init__(self, xlFile, dataRange, uncertRange=None, sheets = None) -> None:


        # check the extension
        extension = os.path.splitext(xlFile)[1]
        supportedExtensions = ['.xls', '.xlsx']
        if extension not in supportedExtensions:
            raise ValueError(f'The file extension is not supported. The supported extension are {supportedExtensions}')

        self.wb = CalamineWorkbook.from_path(xlFile)
        self.sheets = [self.wb.get_sheet_by_name(elem) for elem in self.wb.sheet_names]
        if not sheets is None:
            if not isinstance(sheets, list):
                sheets = [sheets]
            self.sheets = [elem for i,elem in enumerate(self.sheets) if i in sheets]
            
        if not isinstance(dataRange, list):
            dataRange = [dataRange] * len(self.sheets)
        if len(dataRange) != len(self.sheets):
            raise ValueError('The length of the input "dataRange" has to be equal to the length of the intput "sheets"')
        
        if not uncertRange is None:
            if not isinstance(uncertRange, list):
                uncertRange = [uncertRange] * len(self.sheets)
            if len(uncertRange) != len(self.sheets):
                raise ValueError('The length of the input "uncertRange" has to be equal to the length of the intput "sheets"')
        else:
            uncertRange = [None] * len(self.sheets)
            
        
        self.dataRanges = []
        self.uncertRanges = []
        
        self.dataStartCol = []
        self.dataEndCol = []
        self.uncertStartCol = []
        self.uncertEndCol = []    
        for i in range(len(self.sheets)):
            if '-' in dataRange[i]:
                index = dataRange[i].find('-')
                dataStartCol = dataRange[i][0:index]
                dataEndCol = dataRange[i][index + 1:]
            else:
                dataStartCol = dataRange[i]
                dataEndCol = dataStartCol

            if '-' in dataStartCol or '-' in dataEndCol:
                raise ValueError('The data range can only include a singly hyphen (-)')

            if not uncertRange[i] is None:
                if '-' in uncertRange[i]:
                    index = uncertRange[i].find('-')
                    uncertStartCol = uncertRange[i][0:index]
                    uncertEndCol = uncertRange[i][index + 1:]
                else:
                    uncertStartCol = uncertRange[i]
                    uncertEndCol = uncertStartCol

                if '-' in uncertStartCol or '-' in uncertEndCol:
                    raise ValueError('The data range can only include a singly hyphen (-)')
            else:
                uncertStartCol = None
                uncertEndCol = None

            dataStartCol = self.colToIndex(dataStartCol)
            dataEndCol = self.colToIndex(dataEndCol)
            uncertStartCol = self.colToIndex(uncertStartCol)
            uncertEndCol = self.colToIndex(uncertEndCol)
                                            
            self.dataRanges.append(range(dataStartCol, dataEndCol+1) if not dataStartCol is None else None)
            self.uncertRanges.append(range(uncertStartCol, uncertEndCol+1) if not uncertStartCol is None else None)
        
        # read the data
        self.readData()

    def colToIndex(self, col):
        if col is None:
            return None
        if not isinstance(col, str):
            raise ValueError('The coloumn has to be a string')
        num = 0
        for c in col:
            if c in string.ascii_letters:
                num = num * 26 + (ord(c.upper()) - ord('A')) + 1
        return num - 1

    def formatHeaders(self, header):

        out = []
        for head in header:
            if head is None:
                out.append('')
                continue
            # remove symbols and replace with _
            head = re.sub(r'[^\w]', '_', head)
            
            # remove multiple consequtive _
            done = False
            while not done:
                head = head.replace('__', '_')
                if not '__' in head:
                    done = True
                         
            # add "_" to the begining of the name if the first letter is a digit
            if head[0].isnumeric():
                head = '_' + head

            # remove "_" if the last letter is "_"
            if head[-1] == "_" and len(head) != 1:
                head = head[0:-1]

            # add a number to the end of the unit if the unit exists
            if head not in out:
                out.append(head)
            else:
                i, imax, done = 1, 100, False
                while not done and i <= imax:
                    h = head + f'_{i+1}'
                    if h not in out:
                        out.append(h)
                        done = True
                    i += 1
                if not done:
                    raise ValueError(f'The header {head} could not be added to the sheet')
        return out

    def readData(self):
        self.dat = []

        # Looping over the sheets in the data file
        for ii, sh in enumerate(self.sheets):
            
            sheetData = sheet()
            
            workbook = sh.to_python(skip_empty_area=False)
                        
            dataRange = self.dataRanges[ii]
            uncertRange = self.uncertRanges[ii]
            nDataColoumns = len(dataRange)

            
            if not uncertRange is None:
                nUncertColoumns = len(uncertRange)
                if nDataColoumns != nUncertColoumns:
                    raise ValueError('The number of data coloumns does not match the number of uncertanty coloumn')
                        
            ## get the headers and the units
            headers, units, workbook = workbook[0], workbook[1], workbook[2:]

            headers = [headers[i] for i in dataRange]
            units = [units[i] for i in dataRange]

            
            ## read the data
            data = []
            for i in dataRange:
                dat = [elem[i] for elem in workbook]

                indexNan = -1
                lastWasNan = False
                for ii in range(len(dat)):
                    if lastWasNan:
                        if dat[ii] != '':
                            lastWasNan = False
                            indexNan = -1
                    else:
                        if dat[ii] != '':
                            indexNan = ii  
                data.append(dat[0:indexNan+1])

            validData = False
            for dat in data:
                if dat:
                    validData = True
                    break
            if not validData:
                raise ValueError("There is no valid data in the data sheet")
            
            
            nData = max([len(d) for d in data])
            for i, d in enumerate(data):
                if len(d) != nData:
                    data[i] = [np.nan] * nData  

            uncertIsZero = uncertRange is None
            if not uncertRange is None:
                uncert = []
                for i in uncertRange:
                    unc = [elem[i] for elem in workbook]

                    indexNan = -1
                    lastWasNan = False
                    for ii in range(len(unc)):
                        if lastWasNan:
                            if unc[ii] != '':
                                lastWasNan = False
                                indexNan = -1
                        else:
                            if unc[ii] != '':
                                indexNan = ii  
                    uncert.append(unc[0:indexNan+1])
                    validUncert = False
                    for unc in uncert:
                        if unc:
                            validUncert = True
                            break
                    if not validUncert:
                        raise ValueError("There is no valid uncertanty in the data sheet")
                    
            if uncertIsZero:
                n = len(data)
                m = len(data[0])
                uncert = [[0] * m] * n     
              
            nUncert = len(uncert[0])
            for u in uncert:
                if len(u) != nUncert:
                    raise ValueError('The uncertanty coloumns does not have an equal length')



            if nData == nUncert:
                containsCovariace = False
            else:
                if nUncert == nData * nDataColoumns:
                    containsCovariace = True
                else:
                    raise ValueError('The number of rows in the uncertanty does not match the data')
                
            if not containsCovariace:
                variables = []
                for i in range(len(list(dataRange))):
                    dat, u, unc = data[i], units[i], uncert[i]
                    dat = [elem if isinstance(elem, float) or isinstance(elem, int) else np.nan for elem in dat]
                    unc = [elem if isinstance(elem, float) or isinstance(elem, int) else np.nan for elem in unc]
                    variables.append(variable(dat, u, unc))


            else:
                
                ## extract the covariance matricies of each setpoint
                uncert = np.array(uncert).transpose()
                covariances = []
                for i in range(nData):
                    covariances.append(uncert[(i) * nDataColoumns:(i+1) * nDataColoumns , : ])
                covariances = np.array(covariances)
                
                for cov in covariances:
                    for i in range(nDataColoumns):
                        for j in range(nDataColoumns):
                            if cov[i,j] != cov[j,i]:
                                raise ValueError("The covariances has to be symmetric")
                
                ## get the uncertanty as the diagonal of each covariance matrix
                uncert = []
                for i in range(len(list(dataRange))):
                    uncert.append(np.array([np.sqrt(cov[i,i]) for cov in covariances]))
                
                variables = [variable(data[i], units[i], uncert[i]) for i in range(len(list(dataRange)))]
                
                for i in range(nDataColoumns):
                    for j in range(nDataColoumns):
                        if i == j: continue
                        variables[i].addCovariance(variables[j], [float(elem) for elem in covariances[:,i,j]], str(variables[i]._unitObject * variables[j]._unitObject))

            for i in range(len(list(dataRange))):
                setattr(sheetData, headers[i], variables[i])

            

            self.dat.append(sheetData)



class sheet():
    def __init__(self):
        pass

    def printContents(self):
        for key, item in self.__dict__.items():
            if isinstance(item, scalarVariable):
                print(key)

    def __getitem__(self, index):
        sh = sheet()
        for key, item in self.__dict__.items():
            if isinstance(item, scalarVariable):
                setattr(sh, key, item[index])
        return sh

    def __len__(self):
        
        for _, item in self.__dict__.items():
            if isinstance(item, scalarVariable):
                return len(item)
        raise 0

    def append(self, other):
        if not isinstance(other, sheet):
            raise ValueError('You can only append two sheets together')

        selfMeasurements = []
        selfMeasurementNames = []
        otherMeasurements = []
        otherMeasurementNames = []
        
        for key,item in self.__dict__.items():
            if isinstance(item,scalarVariable):
                selfMeasurementNames.append(key)
                selfMeasurements.append(item)

        for key,item in other.__dict__.items():
            if isinstance(item,scalarVariable):
                otherMeasurementNames.append(key)
                otherMeasurements.append(item)

        # Test if all names are the same
        for elem in selfMeasurementNames:
            if elem not in otherMeasurementNames:
                raise ValueError('You can only append sheets with the excact same measurements. The names did not match')

        for elem in otherMeasurementNames:
            if elem not in selfMeasurementNames:
                raise ValueError('You can only append sheets with the excact same measurements. The names did not match')

        # append the measurements from other to self
        for measurement, measurementName in zip(selfMeasurements, selfMeasurementNames):
            index = otherMeasurementNames.index(measurementName)
            measurement.append(otherMeasurements[index])    
        
    def __iter__(self):
        
        variables = []
        for _, item in self.__dict__.items():
            if isinstance(item, scalarVariable):
                variables.append(item)
        
        return iter(variables)
    
    def pop(self, index = -1):
        for key, item in self.__dict__.items():
            if isinstance(item, scalarVariable):
                item.pop(index)

    def __setattr__(self, name: str, value: Any) -> None:
        if not isinstance(value, scalarVariable):
            raise ValueError('You can only set variables as attributes to a sheet')
        
        if not hasattr(value, '__len__'):
            ## the variable is not an arrayVariable. It has to be to be an array variable
            value = arrayVariable(scalarVariables=[value])
        
        self.__dict__[name] = value


